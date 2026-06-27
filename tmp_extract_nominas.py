"""
Dump all nomina Excel files to readable text for reverse-engineering.
Reads every non-empty cell in every row, row by row, across all sheets.
Writes UTF-8 output directly to file.
"""
import openpyxl
import xlrd
import glob
import os
import sys

BASE = r"c:\Users\luisc\Documents\Dataholics\Dataholics Guidelines\proyectos\Calculadoras Sotelo\Reportes"
OUT = r"c:\Users\luisc\Documents\Dataholics\Dataholics Guidelines\proyectos\Calculadoras Sotelo\tmp_nominas_dump_utf8.txt"

files = (
    glob.glob(os.path.join(BASE, "Reporte 1", "Foraneo Chihuahua*.xlsx")) +
    glob.glob(os.path.join(BASE, "Reporte 1", "Foraneo pacifico*.xlsx")) +
    glob.glob(os.path.join(BASE, "Reporte 2", "Nomina O.P Cruce*.xls")) +
    glob.glob(os.path.join(BASE, "Reporte 3", "Nomina O.P Cruce*.xls")) +
    glob.glob(os.path.join(BASE, "Reporte 3", "nomina electrocomponentes*.xlsx")) +
    glob.glob(os.path.join(BASE, "Reporte 3", "operadores locales*.xlsx"))
)

with open(OUT, "w", encoding="utf-8") as out:
    for path in sorted(files):
        fname = os.path.basename(path)
        out.write(f"\n{'='*80}\n")
        out.write(f"FILE: {fname}\n")
        out.write(f"{'='*80}\n")
        try:
            if path.lower().endswith(".xls"):
                wb = xlrd.open_workbook(path)
                for sheet_name in wb.sheet_names():
                    ws = wb.sheet_by_name(sheet_name)
                    out.write(f"\n  --- Sheet: {sheet_name} ---\n")
                    for r in range(ws.nrows):
                        cells = []
                        for c in range(ws.ncols):
                            val = ws.cell_value(r, c)
                            if str(val).strip() != "":
                                col = openpyxl.utils.get_column_letter(c + 1)
                                coord = f"{col}{r + 1}"
                                cells.append(f"[{coord}]={val}")
                        if cells:
                            out.write("  " + " | ".join(cells) + "\n")
            else:
                wb = openpyxl.load_workbook(path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    out.write(f"\n  --- Sheet: {sheet_name} ---\n")
                    for row in ws.iter_rows():
                        cells = []
                        for cell in row:
                            val = cell.value
                            if val is not None and str(val).strip() != "":
                                cells.append(f"[{cell.coordinate}]={val}")
                        if cells:
                            out.write("  " + " | ".join(cells) + "\n")
        except Exception as e:
            out.write(f"  ERROR: {e}\n")

print(f"Done. Written to {OUT}")
