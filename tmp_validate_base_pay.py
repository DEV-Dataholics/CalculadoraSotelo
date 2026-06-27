# =============================================================================
# VALIDACION DE PAGO BASE vs NOMINAS PAGADAS
# Compara el "Pago por KM" (base, sin diesel) de nuestra lógica contra los
# valores reales en las nóminas FCH y PAC semanas 41-45.
#
# HALLAZGO CLAVE: FCH también es km-base, no flat $110/$55.
#   Rate cargado = 0.29333 /km  (≈ $110 / 375km ruta estándar)
#   Rate vacío   = 0.14666 /km  (≈ $55  / 375km ruta estándar)
# =============================================================================

import os
import sys
import openpyxl
import pandas as pd
from datetime import datetime

BASE = r"c:\Users\luisc\Documents\Dataholics\Dataholics Guidelines\proyectos\Calculadoras Sotelo"
REPORTES_1 = os.path.join(BASE, "Reportes", "Reporte 1")
GENESIS_DIR = os.path.join(BASE, "Genesis")

# ---------------------------------------------------------------------------
# Rates
# ---------------------------------------------------------------------------
RATE_LOADED_FCH  = 0.29333   # per km — confirmed from nomina formula
RATE_EMPTY_FCH   = 0.14666   # per km
RATE_LOADED_PAC  = 0.30      # per km
RATE_EMPTY_PAC   = 0.15      # per km

# CVP codes in nómina col N (C/V/PT column):
#   1 = Cargado, 2 = Vacío, 3 = PT
#   4,5,6 = ELP variants (cargado with km deduction: 40, 50, 90)
ELP_DEDUCTIONS = {4: 40, 5: 50, 6: 90}

# Route code → km (from KILOMETRAJE sheet)
R_TABLE = {
    1: 425, 2: 375, 3: 375, 4: 415, 5: 457, 6: 457,
    7: 480, 8: 480, 9: 105, 10: 105, 11: 490, 12: 490,
    13: 680, 14: 680, 15: 1021, 16: 1021, 17: 1330, 18: 1330,
    19: 219, 20: 219, 21: 348, 22: 348, 23: 425, 24: 465,
    25: 538, 26: 538, 27: 415, 28: 465, 29: 465, 30: 425,
    31: 425, 32: 465, 33: 145, 34: 145, 35: 1131,
}

# ---------------------------------------------------------------------------
# Skip sheets that are not individual driver payroll
# ---------------------------------------------------------------------------
SKIP_SHEETS = {'NOMINA', 'KM POR CHOFER', 'KM POR UNIDAD', 'KM POR CLIENTE',
               'KILOMETRAJE', 'UNIDADES', 'FORMULAS', 'Sheet1', 'Sheet2'}

# ---------------------------------------------------------------------------
# Parse one driver sheet — returns dict with extracted data
# ---------------------------------------------------------------------------
def parse_driver_sheet(ws, sheet_name, semana, tipo):
    """
    Reads a driver sheet (data_only=True) and extracts:
    - driver name, unit
    - per-row movement data (coordenada, R, kms, cvpt, actual_pay_row)
    - totals: actual_km_pay_total
    """
    # Row 11: OPERADOR name (col E = index 5), UNIDAD (col E row 12)
    driver = ws['E11'].value or sheet_name
    unit   = ws['E12'].value or ''
    
    movements = []
    # Data rows are 19–33 (FCH 15-row template) or 19–28 (PAC/short 10-row template)
    # Scan all possible rows; stop early if we hit the CRUCES/PERCEPCIONES section
    for r in range(19, 34):
        # Stop if col E contains a non-date string marker (e.g. 'CRUCES')
        e_val = ws.cell(r, 5).value
        if isinstance(e_val, str) and 'CRUCES' in e_val.upper():
            break
        coord  = ws.cell(r, 4).value   # col D: Coordenada
        r_code = ws.cell(r, 7).value   # col G: R code
        km_val = ws.cell(r, 10).value  # col J: KMS (resolved)
        cvpt   = ws.cell(r, 14).value  # col N: C/V/PT code
        pay_q  = ws.cell(r, 17).value  # col Q: PAGO POR KM (resolved)
        folio  = ws.cell(r, 3).value   # col C: Folio

        if coord is None and r_code is None:
            continue  # empty row

        try:
            km_val = float(km_val) if km_val is not None else 0.0
            cvpt   = int(cvpt)     if cvpt   is not None else 0
            pay_q  = float(pay_q)  if pay_q  is not None else 0.0
            r_code = int(r_code)   if r_code is not None else 0
        except (ValueError, TypeError):
            continue

        # Recalculate pay using our rule
        # FCH legs through Pacific destinations use PAC rate (0.30/0.15)
        origin_v = str(ws.cell(r, 8).value or '').upper()
        dest_v   = str(ws.cell(r, 9).value or '').upper()
        PAC_KW = ('OBRG', 'OBREGON', 'MOCHIS', 'GUAMUCHIL', 'NAVOJOA', 'CANANEA',
                  'ETCHO', 'JANOS', 'NOGALES', 'S. RIO COL', 'HERMOSILLO', 'EMPALME', 'BACUM')
        is_pac_leg = tipo == 'PAC' or any(k in origin_v or k in dest_v for k in PAC_KW)
        rate   = RATE_LOADED_PAC  if is_pac_leg else RATE_LOADED_FCH
        rate_e = RATE_EMPTY_PAC   if is_pac_leg else RATE_EMPTY_FCH

        if cvpt == 1:
            our_pay = km_val * rate
        elif cvpt in (2, 3):
            our_pay = km_val * rate_e
        elif cvpt in ELP_DEDUCTIONS:
            our_pay = max(0, km_val - ELP_DEDUCTIONS[cvpt]) * rate
        else:
            our_pay = 0.0

        movements.append({
            'folio':    folio,
            'coord':    str(coord) if coord else '',
            'r_code':   r_code,
            'km':       km_val,
            'cvpt':     cvpt,
            'actual_q': round(pay_q, 2),
            'our_q':    round(our_pay, 2),
            'delta_q':  round(our_pay - pay_q, 2),
        })

    # Locate PAGO POR KM total dynamically (row varies by template length: F33 short, F38 long)
    total_actual = None
    for r in range(30, 45):
        label = ws.cell(r, 3).value  # col C
        if label and 'PAGO POR KM' in str(label).upper():
            val = ws.cell(r, 6).value  # col F
            if val is not None:
                try:
                    total_actual = float(val)
                except (ValueError, TypeError):
                    pass
            break

    # Sum our recalculated per-row pays
    our_total = round(sum(m['our_q'] for m in movements), 2)

    return {
        'semana': semana,
        'tipo':   tipo,
        'driver': str(driver).strip(),
        'unit':   str(unit).strip(),
        'movements': movements,
        'actual_km_pay': round(total_actual, 2) if total_actual is not None else None,
        'our_km_pay':    our_total,
        'delta':         round(our_total - total_actual, 2) if total_actual is not None else None,
    }


# ---------------------------------------------------------------------------
# Process one nomina file
# ---------------------------------------------------------------------------
def process_nomina_file(filepath, semana, tipo):
    results = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ERROR abriendo {os.path.basename(filepath)}: {e}")
        return results

    for name in wb.sheetnames:
        if name.upper() in SKIP_SHEETS:
            continue
        ws = wb[name]
        try:
            rec = parse_driver_sheet(ws, name, semana, tipo)
            if rec['movements']:  # skip blank sheets
                results.append(rec)
        except Exception as e:
            print(f"  ERROR en hoja '{name}': {e}")
    return results


# ---------------------------------------------------------------------------
# Collect all FCH & PAC files
# ---------------------------------------------------------------------------
def find_nomina_files():
    files = []
    for fname in os.listdir(REPORTES_1):
        lower = fname.lower()
        if not fname.endswith('.xlsx'):
            continue
        # Extract semana number
        import re as _re
        m = _re.search(r'#(\d+)', fname)
        if m:
            sem = int(m.group(1))
        else:
            sem = None
        if sem is None:
            continue
        if 'chihuahua' in lower:
            files.append((os.path.join(REPORTES_1, fname), sem, 'FCH'))
        elif 'pacifico' in lower or 'pacífico' in lower:
            files.append((os.path.join(REPORTES_1, fname), sem, 'PAC'))
    return sorted(files, key=lambda x: (x[2], x[1]))


# ---------------------------------------------------------------------------
# Main — run validation and print report
# ---------------------------------------------------------------------------
def main():
    files = find_nomina_files()
    print(f"Archivos encontrados: {len(files)}")
    for f, s, t in files:
        print(f"  Sem{s} {t}: {os.path.basename(f)}")
    print()

    all_records = []
    for fpath, semana, tipo in files:
        print(f"Procesando Sem{semana} {tipo}...")
        recs = process_nomina_file(fpath, semana, tipo)
        all_records.extend(recs)
        print(f"  {len(recs)} choferes procesados")

    if not all_records:
        print("No se encontraron datos.")
        return

    # ------------------------------------------------------------------
    # Summary table
    # ------------------------------------------------------------------
    print()
    print("=" * 90)
    print(f"{'SEM':<5} {'TIPO':<5} {'CHOFER':<35} {'UNIDAD':<8} "
          f"{'REAL $KM':>10} {'CALC $KM':>10} {'DELTA':>9} {'ESTADO'}")
    print("=" * 90)

    match_count    = 0
    mismatch_count = 0
    no_data_count  = 0
    total_delta    = 0.0
    mismatch_rows  = []

    for r in all_records:
        if r['actual_km_pay'] is None:
            status = "SIN DATO"
            no_data_count += 1
        elif abs(r['delta']) <= 1.0:
            status = "✓ OK"
            match_count += 1
        else:
            pct = abs(r['delta']) / r['actual_km_pay'] * 100 if r['actual_km_pay'] else 0
            status = f"✗ {pct:.1f}%"
            mismatch_count += 1
            mismatch_rows.append(r)
            if r['delta'] is not None:
                total_delta += r['delta']

        actual_str = f"${r['actual_km_pay']:,.2f}" if r['actual_km_pay'] is not None else "N/A"
        delta_str  = f"{r['delta']:+,.2f}" if r['delta'] is not None else "N/A"
        print(f"{r['semana']:<5} {r['tipo']:<5} {r['driver'][:34]:<35} {r['unit']:<8} "
              f"{actual_str:>10} ${r['our_km_pay']:>9,.2f} {delta_str:>9} {status}")

    print("=" * 90)
    print(f"Total choferes: {len(all_records)}  |  OK: {match_count}  |  "
          f"Diferencias: {mismatch_count}  |  Sin dato: {no_data_count}")
    if mismatch_count > 0:
        print(f"Delta acumulado en diferencias: ${total_delta:+,.2f}")

    # ------------------------------------------------------------------
    # Drill-down on rows with biggest differences
    # ------------------------------------------------------------------
    if mismatch_rows:
        print()
        print("=" * 90)
        print("DETALLE — MOVIMIENTOS CON DIFERENCIAS (primeros 3 choferes)")
        print("=" * 90)
        for r in sorted(mismatch_rows, key=lambda x: abs(x['delta'] or 0), reverse=True)[:3]:
            print(f"\nSem{r['semana']} {r['tipo']} | {r['driver']} | {r['unit']}")
            print(f"  Real: ${r['actual_km_pay']:,.2f}  Calc: ${r['our_km_pay']:,.2f}  Delta: ${r['delta']:+,.2f}")
            print(f"  {'COORD':<20} {'R':>4} {'KM':>6} {'CVT':>4} {'REAL_Q':>9} {'CALC_Q':>9} {'DELTA_Q':>9}")
            for m in r['movements']:
                if abs(m['delta_q']) > 0.5:
                    flag = " <-- DIFF"
                else:
                    flag = ""
                print(f"  {m['coord'][:19]:<20} {m['r_code']:>4} {m['km']:>6.0f} "
                      f"{m['cvpt']:>4} ${m['actual_q']:>8,.2f} ${m['our_q']:>8,.2f} "
                      f"{m['delta_q']:>+9.2f}{flag}")

    # ------------------------------------------------------------------
    # Key findings summary
    # ------------------------------------------------------------------
    print()
    print("=" * 90)
    print("HALLAZGOS DE VALIDACION")
    print("=" * 90)
    ok_pct = match_count / len(all_records) * 100 if all_records else 0
    print(f"  Precisión base-pay: {ok_pct:.1f}% ({match_count}/{len(all_records)} choferes dentro de $1)")
    print(f"  Tasa FCH aplicada: $0.29333/km cargado | $0.14666/km vacío/PT")
    print(f"  Tasa PAC aplicada: $0.30/km cargado    | $0.15/km vacío/PT")
    print(f"  ELP deductions: cvpt=4 →-40km, cvpt=5 →-50km, cvpt=6 →-90km")
    print()

    # Per-tipo summary
    fch_recs = [r for r in all_records if r['tipo'] == 'FCH' and r['actual_km_pay'] is not None]
    pac_recs = [r for r in all_records if r['tipo'] == 'PAC' and r['actual_km_pay'] is not None]
    if fch_recs:
        fch_ok = sum(1 for r in fch_recs if abs(r['delta']) <= 1.0)
        print(f"  FCH: {fch_ok}/{len(fch_recs)} OK  | "
              f"Delta total: ${sum(r['delta'] for r in fch_recs):+,.2f}")
    if pac_recs:
        pac_ok = sum(1 for r in pac_recs if abs(r['delta']) <= 1.0)
        print(f"  PAC: {pac_ok}/{len(pac_recs)} OK  | "
              f"Delta total: ${sum(r['delta'] for r in pac_recs):+,.2f}")


if __name__ == '__main__':
    main()
